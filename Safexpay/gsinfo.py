import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import getpass
usern=getpass.getuser()


import_file_path = 'C:/Users/'+usern+'/Documents/Cards_safexpay/cards.xlsx'
export_file_path = 'C:/Users/'+usern+'/Documents/Cards_safexpay/cardsreport.xlsx'

def getinfo():
    df=pd.read_excel(import_file_path)
    cardno_list = df['CARDNO'].tolist()
    exp_list = df['EXP_DATE'].tolist()
    cvv_list = df['CVV'].tolist()
    ipin_list = df['IPIN'].tolist()
    name_list = df['NAME'].tolist()
    email_list = df['EMAIL'].tolist()
    phone_list = df['PHONE'].tolist()
    ilength=(len(cardno_list))
    return(cardno_list, exp_list, cvv_list, ipin_list, ilength, name_list, email_list, phone_list)
    
def repheader():
    wb = load_workbook(export_file_path)
    work_sheet = wb.active # Get active sheet
    work_sheet.append(['CARDNO', 'EXP_DATE', 'CVV', 'IPIN', 'AMOUNT', 'TRANSACTION_ID', 'TIME_TAKEN', 'STATUS'])
    wb.save(export_file_path)
    
    
def exwrite(o_cardno_list, o_exp_list, o_cvv_list, o_ipin_list, transaction_id_list, t_time_list, status, pamount):
    wb = load_workbook(export_file_path)
    work_sheet = wb.active # Get active sheet
    work_sheet.append([o_cardno_list, o_exp_list, o_cvv_list, o_ipin_list, pamount, transaction_id_list, t_time_list, status])
    wb.save(export_file_path)
    
def summary(status, pamount, total_time):
    now=datetime.now()
    dt=now.strftime("%d/%m/%Y %H:%M:%S")
    wb = load_workbook(export_file_path)
    work_sheet = wb.active # Get active sheet
    work_sheet.append(['Successfull Payments', status.count("Success")])
    work_sheet.append(['Failed Payments', status.count("Fail")])
    work_sheet.append(['Total Payments', status.count("Success")+status.count("Fail")])
    work_sheet.append(['Total Amount Paid', int(pamount)*int(status.count("Success"))])
    work_sheet.append(['Total Time Taken', total_time])
    work_sheet.append(['Date, Time', dt])
    wb.save(export_file_path)  
