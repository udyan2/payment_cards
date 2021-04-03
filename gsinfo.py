import pandas as pd
from openpyxl import load_workbook
from datetime import datetime


def getinfo():
    import_file_path = 'C:/Users/AVITA/Documents/Cards/cards.xlsx'
    df=pd.read_excel(import_file_path)
    cardno_list = df['CARDNO'].tolist()
    exp_list = df['EXP_DATE'].tolist()
    cvv_list = df['CVV'].tolist()
    ipin_list = df['IPIN'].tolist()
    ilength=(len(cardno_list))
    return(cardno_list, exp_list, cvv_list, ipin_list, ilength)

# def sendinfo(o_cardno_list, o_exp_list, o_cvv_list, transaction_id_list, t_time_list, status, total_time, pamount):
#     now=datetime.now()
#     dt=now.strftime("%d/%m/%Y %H:%M:%S")
#     import_file_path = 'C:/Users/AVITA/Documents/Cards/cardsreport.xlsx'
#     sdf=pd.read_excel(import_file_path)
#     sdf["CARDNO"] = o_cardno_list
#     sdf.to_excel("import_file_path",index=False)
#     sdf["EXP_DATE"] = o_exp_list
#     sdf.to_excel("import_file_path",index=False)
#     sdf["CVV"] = o_cvv_list
#     sdf.to_excel("import_file_path",index=False)
#     sdf["TRANSACTION_ID"] = transaction_id_list
#     sdf.to_excel("import_file_path",index=False)
#     sdf["TIME_TAKEN"] = t_time_list
#     sdf.to_excel("import_file_path",index=False)
#     sdf["STATUS"] = status
#     sdf.to_excel("import_file_path",index=False)
#     wb = load_workbook('import_file_path')
#     work_sheet = wb.active # Get active sheet
#     work_sheet.append(['Successfull Payments', status.count("Success")])
#     work_sheet.append(['Failed Payments', status.count("Fail")])
#     work_sheet.append(['Total Payments', status.count("Success")+status.count("Fail")])
#     work_sheet.append(['Total Amount Paid', int(pamount)*int(status.count("Success"))])
#     work_sheet.append(['Total Time Taken', total_time])
#     work_sheet.append(['Date, Time', dt])
#     wb.save('import_file_path')
    
def repheader():
    import_file_path = 'C:/Users/AVITA/Documents/Cards/cardsreport.xlsx'
    wb = load_workbook(import_file_path)
    work_sheet = wb.active # Get active sheet
    work_sheet.append(['CARDNO', 'EXP_DATE', 'CVV', 'TRANSACTION_ID', 'AMOUNT','TIME_TAKEN', 'STATUS'])
    wb.save(import_file_path)
    
    
def exwrite(o_cardno_list, o_exp_list, o_cvv_list, transaction_id_list, t_time_list, status, total_time, pamount):
    import_file_path = 'C:/Users/AVITA/Documents/Cards/cardsreport.xlsx'
    wb = load_workbook(import_file_path)
    work_sheet = wb.active # Get active sheet
    work_sheet.append([o_cardno_list, o_exp_list, o_cvv_list, transaction_id_list, pamount, t_time_list, status])
    wb.save(import_file_path)
    
def summary(status, pamount, total_time):
    now=datetime.now()
    dt=now.strftime("%d/%m/%Y %H:%M:%S")
    import_file_path = 'C:/Users/AVITA/Documents/Cards/cardsreport.xlsx'
    wb = load_workbook(import_file_path)
    work_sheet = wb.active # Get active sheet
    work_sheet.append(['Successfull Payments', status.count("Success")])
    work_sheet.append(['Failed Payments', status.count("Fail")])
    work_sheet.append(['Total Payments', status.count("Success")+status.count("Fail")])
    work_sheet.append(['Total Amount Paid', int(pamount)*int(status.count("Success"))])
    work_sheet.append(['Total Time Taken', total_time])
    work_sheet.append(['Date, Time', dt])
    wb.save('import_file_path')  
